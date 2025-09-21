#!/usr/bin/env python3
"""
Script para generar tabla Excel profesional con resultados del proyecto HPC
Crea un archivo .xlsx con formato, colores y múltiples hojas organizadas
"""

import pandas as pd
import numpy as np

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("⚠️ openpyxl no disponible. Se generará CSV mejorado.")

# Datos del proyecto HPC
data = {
    'Tamaño': [10, 10, 10, 10, 10, 100, 100, 100, 100, 100, 500, 500, 500, 500, 500, 1000, 1000, 2000],
    'Hilos': [1, 2, 4, 8, 16, 1, 2, 4, 8, 16, 1, 2, 4, 8, 16, 1, 8, 4],
    'T_Secuencial': [0.0000, 0.0000, 0.0000, 0.0000, 0.0000, 0.0080, 0.0050, 0.0050, 0.0050, 0.0060, 
                     1.0250, 0.6570, 0.7160, 0.6360, 0.6270, 6.7540, 7.0030, 91.7810],
    'T_Paralelo_Avg': [0.0003, 0.0007, 0.0007, 0.0007, 0.0017, 0.0067, 0.0040, 0.0030, 0.0033, 0.0033,
                       0.8827, 0.5540, 0.3250, 0.2340, 0.2720, 8.2160, 2.4113, 49.8180],
    'T_Paralelo_Min': [0.0000, 0.0000, 0.0000, 0.0000, 0.0010, 0.0060, 0.0030, 0.0030, 0.0030, 0.0030,
                       0.8550, 0.5400, 0.3230, 0.2130, 0.2270, 7.9950, 2.1940, 45.2740],
    'T_Paralelo_Max': [0.0010, 0.0010, 0.0020, 0.0010, 0.0020, 0.0080, 0.0050, 0.0030, 0.0040, 0.0040,
                       0.9010, 0.5690, 0.3280, 0.2510, 0.3090, 8.4880, 2.6220, 54.3620],
    'Speedup': [0.00, 0.00, 0.00, 0.00, 0.00, 1.20, 1.25, 1.67, 1.50, 1.80,
                1.16, 1.19, 2.20, 2.72, 2.31, 0.82, 2.90, 1.84],
    'Eficiencia': [0.0, 0.0, 0.0, 0.0, 0.0, 120.0, 62.5, 41.7, 18.8, 11.2,
                   116.1, 59.3, 55.1, 34.0, 14.4, 82.2, 36.3, 46.1],
    'MFLOPS': [0, 0, 0, 0, 0, 298, 500, 667, 606, 606, 283, 451, 769, 1068, 918, 244, 829, 321],
    'Estado': ['No Apto', 'No Apto', 'No Apto', 'No Apto', 'No Apto', 'Baseline', 'Moderado', 'Bueno', 
               'Aceptable', 'MEJOR', 'Baseline', 'Moderado', 'Excelente', 'ÓPTIMO ⭐', 'Muy Bueno', 
               'Baseline', 'MÁXIMO ⭐⭐', 'Bueno']
}

df = pd.DataFrame(data)

def crear_excel_profesional():
    """Crear archivo Excel con formato profesional"""
    if not OPENPYXL_AVAILABLE:
        print("Instalando openpyxl...")
        import subprocess
        subprocess.run(["pip", "install", "openpyxl"], check=True)
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.utils.dataframe import dataframe_to_rows
    
    # Crear libro de trabajo
    wb = Workbook()
    wb.remove(wb.active)  # Remover hoja por defecto
    
    # === HOJA 1: RESUMEN EJECUTIVO ===
    ws_resumen = wb.create_sheet("📊 Resumen Ejecutivo")
    
    # Título principal
    ws_resumen['A1'] = "🚀 PROYECTO HPC - MULTIPLICACIÓN DE MATRICES CON HILOS"
    ws_resumen['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws_resumen['A1'].fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
    ws_resumen.merge_cells('A1:I1')
    
    ws_resumen['A2'] = "📈 ANÁLISIS DE SPEEDUP Y RENDIMIENTO"
    ws_resumen['A2'].font = Font(size=14, bold=True, color="FFFFFF")
    ws_resumen['A2'].fill = PatternFill(start_color="A23B72", end_color="A23B72", fill_type="solid")
    ws_resumen.merge_cells('A2:I2')
    
    # Métricas principales
    ws_resumen['A4'] = "🏆 SPEEDUP MÁXIMO ALCANZADO:"
    ws_resumen['D4'] = "2.90x"
    ws_resumen['D4'].font = Font(size=14, bold=True, color="FF0000")
    
    ws_resumen['A5'] = "⭐ CONFIGURACIÓN ÓPTIMA:"
    ws_resumen['D5'] = "1000×1000 con 8 hilos"
    
    ws_resumen['A6'] = "📊 CONFIGURACIONES PROBADAS:"
    ws_resumen['D6'] = "17 combinaciones"
    
    ws_resumen['A7'] = "✅ PRECISIÓN:"
    ws_resumen['D7'] = "100% verificado"
    
    # === HOJA 2: DATOS COMPLETOS ===
    ws_datos = wb.create_sheet("📋 Datos Completos")
    
    # Encabezados
    headers = ['Tamaño', 'Hilos', 'T_Secuencial(s)', 'T_Paralelo_Avg(s)', 'T_Paralelo_Min(s)', 
               'T_Paralelo_Max(s)', 'Speedup', 'Eficiencia(%)', 'MFLOPS', 'Estado']
    
    for col, header in enumerate(headers, 1):
        cell = ws_datos.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Datos
    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws_datos.cell(row=row_idx, column=col_idx, value=value)
            
            # Colorear según speedup
            if col_idx == 7:  # Columna Speedup
                if value >= 2.5:
                    cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                elif value >= 2.0:
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                elif value >= 1.5:
                    cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    
    # === HOJA 3: TOP CONFIGURACIONES ===
    ws_top = wb.create_sheet("🏆 Top Configuraciones")
    
    # Datos del top 5
    top_configs = [
        ["🥇 1º", "1000×1000, 8h", "2.90x", "36.3%", "829", "MÁXIMO"],
        ["🥈 2º", "500×500, 8h", "2.72x", "34.0%", "1068", "ÓPTIMO"],
        ["🥉 3º", "500×500, 16h", "2.31x", "14.4%", "918", "MUY BUENO"],
        ["4º", "500×500, 4h", "2.20x", "55.1%", "769", "BALANCEADO"],
        ["5º", "2000×2000, 4h", "1.84x", "46.1%", "321", "BUENO"]
    ]
    
    headers_top = ["Ranking", "Configuración", "Speedup", "Eficiencia", "MFLOPS", "Estado"]
    
    for col, header in enumerate(headers_top, 1):
        cell = ws_top.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    
    for row_idx, config in enumerate(top_configs, 2):
        for col_idx, value in enumerate(config, 1):
            cell = ws_top.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 2:  # Primera fila (campeón)
                cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            elif row_idx == 3:  # Segunda fila
                cell.fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
            elif row_idx == 4:  # Tercera fila
                cell.fill = PatternFill(start_color="CD7F32", end_color="CD7F32", fill_type="solid")
    
    # === HOJA 4: ANÁLISIS POR TAMAÑO ===
    ws_analisis = wb.create_sheet("📊 Análisis por Tamaño")
    
    # Análisis por tamaño
    sizes = [10, 100, 500, 1000, 2000]
    
    row_start = 1
    for size in sizes:
        data_size = df[df['Tamaño'] == size]
        if not data_size.empty:
            # Título de sección
            ws_analisis.cell(row=row_start, column=1, value=f"MATRIZ {size}×{size}").font = Font(bold=True, size=12)
            ws_analisis.cell(row=row_start, column=1).fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            ws_analisis.merge_cells(f'A{row_start}:I{row_start}')
            
            # Headers
            row_start += 1
            for col, header in enumerate(['Hilos', 'Speedup', 'Eficiencia', 'MFLOPS', 'Estado'], 1):
                ws_analisis.cell(row=row_start, column=col, value=header).font = Font(bold=True)
            
            # Datos
            row_start += 1
            for _, row in data_size.iterrows():
                ws_analisis.cell(row=row_start, column=1, value=row['Hilos'])
                ws_analisis.cell(row=row_start, column=2, value=row['Speedup'])
                ws_analisis.cell(row=row_start, column=3, value=f"{row['Eficiencia']:.1f}%")
                ws_analisis.cell(row=row_start, column=4, value=row['MFLOPS'])
                ws_analisis.cell(row=row_start, column=5, value=row['Estado'])
                row_start += 1
            
            row_start += 2  # Espacio entre secciones
    
    # Ajustar ancho de columnas
    for ws in wb.worksheets:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Guardar archivo
    filename = "RESULTADOS_HPC_PROFESIONAL.xlsx"
    wb.save(filename)
    print(f"✅ Archivo Excel generado: {filename}")
    return filename

def crear_tabla_comparativa_csv():
    """Crear CSV optimizado para Excel"""
    
    # Crear tabla por tamaños
    sizes = [10, 100, 500, 1000, 2000]
    rows = []
    
    # Encabezado
    rows.append(["PROYECTO HPC - ANÁLISIS DE SPEEDUP", "", "", "", "", "", "", "", ""])
    rows.append(["Multiplicación de Matrices con Hilos", "", "", "", "", "", "", "", ""])
    rows.append(["", "", "", "", "", "", "", "", ""])
    
    for size in sizes:
        data_size = df[df['Tamaño'] == size]
        if not data_size.empty:
            rows.append([f"MATRIZ {size}×{size}", "", "", "", "", "", "", "", ""])
            rows.append(["Hilos", "T_Secuencial(s)", "T_Paralelo(s)", "Speedup", "Eficiencia(%)", "MFLOPS", "Estado", "Recomendación", ""])
            
            for _, row in data_size.iterrows():
                recomendacion = "SÍ" if row['Speedup'] >= 2.0 else "CONDICIONAL" if row['Speedup'] >= 1.5 else "NO"
                rows.append([
                    row['Hilos'],
                    f"{row['T_Secuencial']:.4f}",
                    f"{row['T_Paralelo_Avg']:.4f}",
                    f"{row['Speedup']:.2f}",
                    f"{row['Eficiencia']:.1f}",
                    row['MFLOPS'],
                    row['Estado'],
                    recomendacion,
                    ""
                ])
            
            # Mejor configuración
            best_idx = data_size['Speedup'].idxmax()
            best = data_size.loc[best_idx]
            rows.append(["MEJOR:", f"{best['Hilos']} hilos", f"{best['Speedup']:.2f}x speedup", "", "", "", "", "", ""])
            rows.append(["", "", "", "", "", "", "", "", ""])
    
    # Resumen final
    rows.append(["RESUMEN EJECUTIVO", "", "", "", "", "", "", "", ""])
    rows.append(["Speedup Máximo:", "2.90x", "(1000×1000, 8h)", "", "", "", "", "", ""])
    rows.append(["Speedup Promedio:", f"{df['Speedup'].mean():.2f}x", "", "", "", "", "", "", ""])
    rows.append(["Configuraciones Exitosas:", f"{len(df[df['Speedup'] >= 1.0])}/17", "88%", "", "", "", "", "", ""])
    rows.append(["Mejor MFLOPS:", "1068", "(500×500, 8h)", "", "", "", "", "", ""])
    
    # Guardar CSV
    filename = "TABLA_COMPARATIVA_EXCEL.csv"
    with open(filename, 'w', newline='', encoding='utf-8') as f:
        import csv
        writer = csv.writer(f)
        writer.writerows(rows)
    
    print(f"✅ Tabla CSV para Excel generada: {filename}")
    return filename

if __name__ == "__main__":
    print("🚀 Generando tablas Excel profesionales...")
    print("="*50)
    
    # Generar CSV optimizado
    csv_file = crear_tabla_comparativa_csv()
    
    # Intentar generar Excel con formato
    try:
        excel_file = crear_excel_profesional()
        print(f"\n📊 ARCHIVOS GENERADOS:")
        print(f"   • {excel_file} (Excel con formato)")
        print(f"   • {csv_file} (CSV para Excel)")
        print(f"   • TABLA_EXCEL_RESULTADOS.csv (CSV detallado)")
    except Exception as e:
        print(f"⚠️ Error generando Excel: {e}")
        print(f"📄 CSV generado: {csv_file}")
    
    print("\n🎯 INSTRUCCIONES:")
    print("1. Abrir cualquier archivo .csv en Excel")
    print("2. Usar 'Datos > Texto en columnas' si es necesario")
    print("3. Aplicar formato de tabla (Ctrl+T)")
    print("4. Agregar colores y gráficos según preferencia")
    
    print("\n✅ ¡Tablas listas para análisis profesional!")
